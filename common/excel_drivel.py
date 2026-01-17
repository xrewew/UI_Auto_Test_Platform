import os
import pathlib
import shutil
import threading
import time

from openpyxl.styles import PatternFill, Font

from conf.setting import DIR_PATH
from logging_conf.logging_config import logs
import traceback
import openpyxl
from common.Webkey import Webkey

#读取excel文件，获取指定的sheet页
# excel_PATH = pathlib.Path(__file__).parents[0].resolve() / 'demo.xlsx'
# excel = openpyxl.load_workbook(str(excel_PATH))
# # sheet = excel['Sheet1'] # 考虑到可能有多个sheet页的用例需要执行，所以需要获取所有sheet页
# print(excel)
#参数解析方法
def arguments(data):
    temp_data = {}
    # 参数要解析，需要基于参数的分号分割数据，再基于等号进行key和value的值的准备，从而实现最终的字典格式
    if data:
        str_temp = data.split(';') #基于分号分割后生成list列表
        for temp in str_temp:
            t = temp.split('=',1) #基于等号再分割一次
            temp_data[t[0]] = t[1] # 以等号左边的字符串为key，等号右边的字符串为value，存储到字典中
    return temp_data

#断言成功后进行单元格的修饰，若PASS则加粗表浅绿，若fail则加粗表浅红
def pass_(cell): #传入一个单元格子
    """
    断言成功后，将单元格的字体加粗并设置为浅绿背景。
    """
    cell.value = 'PASS'
    #设置单元格的样式以及颜色
    cell.fill = PatternFill(patternType='solid', fgColor='AACF91') #单元格为绿色
    cell.font = Font(bold=True) #字体加粗

def fail_(cell):
    """
    断言失败后，将单元格的字体加粗并设置为浅红背景。
    """
    cell.value = 'FAIL'
    #设置单元格的样式以及颜色
    cell.fill = PatternFill(patternType='solid', fgColor='FFCF91') #单元格为红色
    cell.font = Font(bold=True) #字体加粗

#测试用例集合
case = []
sum_files = 0 #记录测试用例文件的数量
def find_testcase():
    """
    查找所有测试用例。
    """
    logs.info('开始查找testcase文件夹下的所有测试用例..............................')
    global case
    global sum_files
    #os.walk方法可以识别文件夹中的内容，包含文件名称与后缀名，子文件夹，以及子文件夹的内容
    for path,dir,files in os.walk(os.path.join(DIR_PATH, 'testcase')):
        for file in files:
            if file.endswith('.xlsx'): #筛选出所有的excel文件
                if '_history' in file:
                    logs.info('查询到历史已执行过的测试文件：'+file+' 已被跳过')
                else:
                    sum_files += 1 # 统计测试用例文件的数量
                    case.append(os.path.join(path, file)) #将当前路径与文件名称拼接起来，添加到测试用例集合中
    if not case:
        logs.error('testcase文件夹下没有任何新添加的测试用例')
        return case
    else:
        logs.info(f'查询到{sum_files}个测试用例文件，分别为：{case},开始执行测试用例')
        return case

#记录失败成功的用例变量
pass_count = 0 #记录成功的用例数量
fail_count = 0 #记录失败的用例数量
fail_case = []
fail_file = None

#创建进程锁，用于在多线程环境下对共享资源进行同步访问
lock = threading.Lock()

def excel_run(testcase):
    """
    从excel文件中读取测试用例，并执行这些用例。
    """
    global pass_count, fail_count, pass_count, fail_case, fail_file
    try:
        excel_path = testcase # 获取测试用例集合中的所有元素，即所有测试用例的路径
        fail_file = excel_path
        logs.info(f"正在加载Excel文件: {excel_path}")
        excel = openpyxl.load_workbook(excel_path)
        for name in excel.sheetnames: #遍历所有sheet页
            sheet = excel[name]
            logs.info(f'当前sheet页名称为：{sheet.title}')
            for value in sheet.values:
                #读取测试用例的正文内容
                if type(value[0]) is int: #基于序号的数据类型来判断是否进入正文
                    logs.info(f"正在执行的操作为：{value[3]}<UNK>")
                # 参数的解析：将参数解析为字典格式。方便直接传入到函数之中。
                    test_data = arguments(value[2]) #获取每一行的操作参数
                    """
                     所有的操作分为如下不同类型：
                                1. 实例化driver对象
                                2. 基于driver对象进行的常规操作行为
                                3. 断言，因为需要对excel文件进行写入
                    """
                    if value[1] == 'open_browser':
                        driver = Webkey(**test_data)
                    elif 'assert' in value[1]: #所有需要断言的测试用例都要以assert开头
                        status = getattr(driver, value[1])(**test_data,expected =value[4])
                        #通过断言方法生成一个结果，基于结果来判断流程的结果是成功还是失败
                        if status:
                            pass_(sheet.cell(row=value[0] + 2, column=6))
                            with lock:
                                pass_count += 1
                        else:
                            fail_(sheet.cell(row=value[0] + 2, column=6))
                            with lock:
                                fail_count += 1
                            fail_case.append(excel_path+'下的：'+name+'第'+str(value[0]+2)+'行测试序号为：'+str(value[0])+'的测试用例') #用来展示失败用例的位置
                        excel.save(excel_path)
                    else:
                        getattr(driver, value[1])(**test_data)

        # 修改文件名在执行完成后，将测试用例文件后面添加_时间_history后缀名
        #当前时间
        current_time = time.strftime("%Y%m%d%H%M%S", time.localtime())
        os.rename(excel_path, excel_path.replace('.xlsx', f'_{current_time}_history.xlsx'))
        logs.info(f'测试用例文件：{excel_path}执行完成，已添加_history后缀名')
        # 关闭excel
        excel.close()
    except:
        #记录出现错误的测试用例文件
        logs.error(f"在执行测试用例文件‘{fail_file}’时发生错误: {traceback.format_exc()}")

# 修改历史执行过的测试用例文件的文件名，将第一个_后面的名字去除。仅用于代码测试便利
def change_excel():
    for path,dir,files in os.walk(os.path.join(DIR_PATH, 'testcase')):
        for file in files:
            if file.endswith('.xlsx'): #筛选出所有的excel文件
                if '_history' in file:
                    logs.info('查询到历史已执行过的测试文件：'+file+'即将修改为原始文件名')
                    # 修复：正确提取原始文件名（保留第一个下划线前的部分 + .xlsx扩展名）
                    base_name = os.path.splitext(file)[0]  # 去掉扩展名
                    original_name = base_name.split('_')[0] + '.xlsx'  # 取第一个下划线前的部分 + 扩展名

                    old_path = os.path.join(path, file)
                    new_path = os.path.join(path, original_name)

                    try:
                        os.rename(old_path, new_path) # 重命名文件（如果目标文件已存在，会抛出FileExistsError）
                        logs.info(f'修改成功，文件名为：{original_name}')
                    except FileExistsError as e:
                        logs.error(f'重命名文件 {file} 失败：{str(e)}')
                else:
                    logs.info('没有检测到历史文件，无需修改，文件列表为'+file+'<UNK>')

def sum_pass_fail():
    """
    计算并返回成功用例数和失败用例数的总和。
    """
    # 打印执行结果
    logs.info(f"测试执行完成，成功用例数: {pass_count}, 失败用例数: {fail_count}")
    for filecase in fail_case:
        logs.info(f'失败用例的位置：{filecase}')

if __name__ == '__main__':
    change_excel()
